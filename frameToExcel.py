from PIL import Image
from openpyxl.styles import PatternFill


def frameToExcel(path, ws, i):
    image = Image.open(path)

    for y in range(1, 89):
        for x in range(1, 119):

            image = image.convert("RGB")

            rgbValue = image.getpixel((x, y))
            hex = '%02x%02x%02x' % rgbValue
            colorFill = PatternFill(start_color=hex, end_color=hex, fill_type='solid')
            ws.cell(y + (i * 100), x).fill = colorFill

    return ws
